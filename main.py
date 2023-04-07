import os
import sys
from typing import Optional, Tuple

from openpyxl.workbook import Workbook

from dataset.constants import EXCEL_FILE_NAME, FOLDER_DIRECTORY
from dataset.config import Configurable as WorkbookSelector
from dataset.datasetclass import Dataset
from dataset.statmeasures import Numeric
from dataset.export import write_dataset_to_worksheet, write_columns_to_worksheet
from ui.selector import Selector, SelectionDisplay
from ui.plotting import get_valid_plot_type, plot_data, plot_compared_data
from ui.functions import (
    create_new_directory,
    get_valid_filename_input, get_workbook_mapping, label_workbooks, get_valid_worksheet_name,
    get_formatted_statistical_function_list, get_statistical_measure_of_region, get_user_decision,
    get_valid_column_name, get_valid_row_number, get_dataset_or_stat_kwargs, valid_column_name,
)

os.chdir(FOLDER_DIRECTORY)
dataset = Dataset(EXCEL_FILE_NAME, "Logan's Dam Water Quality")
unsaved_workbooks = {}

FIRST_COLUMN = dataset.column_names[0]


def get_all_current_workbook_names() -> list:
    # Both unsaved_workbooks and get_workbook_mapping are variable
    return list(unsaved_workbooks) + list(get_workbook_mapping())


def spreadsheet_menu():
    while 1:
        match workbook_options.run():
            case 1:
                if not os.path.exists("workbooks"):
                    create_new_directory("workbooks")
                all_current_workbook_names = get_all_current_workbook_names()
                workbook_name = get_valid_filename_input(existing_filenames=all_current_workbook_names)
                if type(workbook_name) is str:
                    # Should be string unless the loop is broken
                    unsaved_workbooks[workbook_name] = workbook = Workbook()
                    print(f"Created new workbook titled {workbook_name!r}. "
                          f"Save the workbook as a file for later use.")
                    if current_workbook.value is None:
                        print(f"Automatically selected {workbook_name!r} as the current workbook "
                              f"since there is no other selected workbook.")
                        current_workbook.register_value(workbook)
            case 2:
                for workbook_name, workbook in (get_workbook_mapping() | unsaved_workbooks).items():
                    if workbook == current_workbook.value:
                        print(repr(workbook_name), "is currently the selected workbook.")
                        break
                else:
                    print("No workbook is currently selected.")
                all_current_workbook_names = get_all_current_workbook_names()
                if all_current_workbook_names:
                    labelled_workbooks = label_workbooks(unsaved_workbooks, get_workbook_mapping())
                    formatted_labelled_workbooks = [
                        workbook_name + " " + status for workbook_name, status in labelled_workbooks
                    ]
                    workbook_selector = Selector(formatted_labelled_workbooks, notify_option=False)
                    user_input = workbook_selector.run()
                    if type(user_input) is int:
                        selected_workbook_index = user_input - 1
                        selected_workbook_name = labelled_workbooks[selected_workbook_index][0]
                        for workbook_name, workbook in (get_workbook_mapping() | unsaved_workbooks).items():
                            if workbook_name == selected_workbook_name:
                                current_workbook.register_value(workbook)
                                print(f"The current workbook is now {workbook_name!r}.")
                                break
                else:
                    print("There are no existing workbooks. Create a new one.")
            case 3:
                workbook: Workbook = current_workbook.value
                if workbook is not None:
                    new_sheet_name = get_valid_worksheet_name(existing_sheetnames=workbook.sheetnames)
                    workbook.create_sheet(new_sheet_name)
                    # assert workbook.active.title == new_sheet_name
                    print(f"Created new worksheet named {new_sheet_name!r} for file {current_workbook.name}")
                else:
                    print("You have not selected any workbook. Select a workbook before creating a new worksheet")
            case 4:
                workbook: Workbook = current_workbook.value
                if workbook is not None:
                    # current_worksheet.register_value()
                    worksheet_selector = Selector(workbook.sheetnames)
                    if (selected_worksheet := worksheet_selector.run()) is not None:
                        workbook.active = selected_worksheet
                        for workbook_name, wb in (get_workbook_mapping() | unsaved_workbooks).items():
                            if wb == current_workbook.value:
                                current_workbook_name = workbook_name
                                break
                        else:
                            # Should never happen; line implemented to make PyCharm happy
                            current_workbook_name = None
                        print(repr(selected_worksheet), f"is the active worksheet for {current_workbook_name!r}.")
                elif workbook is None:
                    print("You have not selected any workbook. Select a workbook before selecting a worksheet")
                elif not workbook.sheetnames:
                    # This line should never be executed, right? (sheet created by default)
                    print("You have not created any worksheets. Create a new worksheet before selecting one.")
            case 5:
                for workbook_name, workbook in (get_workbook_mapping() | unsaved_workbooks).items():
                    if workbook == current_workbook.value:
                        print(repr(workbook_name))
                        break
                else:
                    print("No workbook is currently selected.")
            case 6:
                if current_workbook.value is not None:
                    print(repr(current_workbook.value.active.title))
                else:
                    print("No worksheet is currently selected because no workbook is currently selected.")
            case 7:
                if current_workbook.value is not None:
                    old_worksheet_name = current_workbook.value.active.title
                    new_worksheet_name = get_valid_worksheet_name(existing_sheetnames=current_workbook.value.sheetnames)
                    current_workbook.value.active.title = new_worksheet_name
                    print(f"\nChanged the name of worksheet {old_worksheet_name!r} to {new_worksheet_name!r}.\n")
                else:
                    print("No active worksheet can be accessed because no workbook is currently selected.")
            case _:
                return


def data_console_menu():
    while 1:
        match console_data_options.run():
            case 1:
                column_name = get_valid_column_name(dataset.column_names)
                if column_name is not None:
                    print(dataset[column_name])
            case 2:
                if (row_number := get_valid_row_number(len(dataset))) is not None:
                    print(dataset[row_number])
            case 3:
                print()
                print(dataset)
            case 4:
                statistical_measures_submenu()
            case 5:
                column_name = get_valid_column_name(dataset.column_names)
                if outliers := dataset.get_outlier_string(column_name):
                    print()
                    print(outliers)
                else:
                    print(f"\nNo outliers found in the column {column_name!r}.")
            case 6:
                print(dataset.get_column_dtypes())
            case 7:
                column_name = get_valid_column_name(dataset.column_names)
                print(dataset[column_name].get_statistical_summary_string(get_formatted_statistical_function_list()))
            case 8:
                print()
                print(dataset.schema)
            case _:
                return


def statistical_measures_submenu(return_data: bool = False) -> Optional[Numeric | Tuple[list, list]]:
    stat_index = statistical_measure_selector.run()
    if type(stat_index) is int:
        statistic = get_formatted_statistical_function_list()[stat_index - 1].lower()
    else:
        return

    enter_kwargs = get_user_decision(
        "Would you like to specify the parameters for the statistics? Type 'y' for yes and 'n' for no "
        "(if no, the default options will be used): ",
        "Please re-enter either 'y' or 'n'"
    )
    kwargs = get_dataset_or_stat_kwargs() if enter_kwargs else {}
    if (data := get_statistical_measure_of_region(dataset)) is None:
        return
    # Polymorphism (data can either be a Dataset or a _DatasetArrayColumnView)
    statistical_data = data.statistic(statistic, **kwargs)
    if return_data:
        return statistical_data

    if type(data) is Dataset:
        cut_data = get_user_decision("Would you like the data to be cut? Type 'y' for yes and 'n' for no: ",
                                     "Please re-enter either 'y' or 'n'")
        print(Dataset.get_statistic_string(*statistical_data, statistic, cut_data=cut_data))
    else:
        print(f"\nThe {statistic} of the column {data.name!r} is {statistical_data}.")


def data_plotting_menu():
    while 1:
        match plot_data_selector.run():
            case 1:
                column_name = get_valid_column_name(dataset.column_names, exclude_first=True)
                print()
                plot_type = get_valid_plot_type()
                plot_data(plot_type, dataset[FIRST_COLUMN], dataset[column_name], FIRST_COLUMN, column_name)
            case 2:
                selected_columns = []
                # Exclude the "date" column
                column_names = dataset.column_names[1:]
                print("\nThe columns and their respective number in the dataset are:", SelectionDisplay(column_names))
                column_identifier = input("Enter the name or position of the first column you would like to graph: ")
                while column_identifier:
                    if column_name := valid_column_name(column_identifier, column_names):
                        selected_columns.append(column_name)
                        break
                    else:
                        print("Please re-enter an appropriate column name or column position.")
                    column_identifier = input(
                        "Enter the name or position of the first column you would like to graph: "
                    )
                column_identifier = input("Enter the name or position of the second column you would like to graph: ")
                while column_identifier:
                    if column_name := valid_column_name(column_identifier, column_names):
                        selected_columns.append(column_name)
                        break
                    else:
                        print("Please re-enter an appropriate column name or column position.")
                    column_identifier = input(
                        "Enter the name or position of the second column you would like to graph: "
                    )
                print()
                if (plot_type := get_valid_plot_type()) is not None:
                    col, other_col = selected_columns
                    plot_compared_data(plot_type, dataset[FIRST_COLUMN], dataset[col],
                                       dataset[other_col], col, other_col)
            case _:
                return


def export_data_menu():
    if current_workbook.value is None:
        print("Select a workbook or create a new workbook using the 'Modify Excel files' menu "
              "before using this menu.")
        return
    match export_data_selector.run():
        case 1:
            enter_kwargs = get_user_decision(
                "Would you like to specify the parameters for the statistics? Type 'y' for yes and 'n' for no "
                "(if no, the default options will be used): ",
                "Please re-enter either 'y' or 'n'"
            )
            kwargs = get_dataset_or_stat_kwargs() if enter_kwargs else {}
            write_dataset_to_worksheet(current_workbook.value.active, dataset.reformat(**kwargs))
            print(f"Wrote data from the modified dataset to worksheet {current_workbook.value.active.title!r}.")
        case 2:
            stat_index = statistical_measure_selector.run()
            if type(stat_index) is int:
                statistic = get_formatted_statistical_function_list()[stat_index - 1].lower()
            else:
                return
            enter_kwargs = get_user_decision(
                "Would you like to specify the parameters for the statistics? Type 'y' for yes and 'n' for no "
                "(if no, the default options will be used): ",
                "Please re-enter either 'y' or 'n'"
            )
            kwargs = get_dataset_or_stat_kwargs() if enter_kwargs else {}
            column_names, statistical_data = dataset.statistic(statistic, **kwargs)
            write_columns_to_worksheet(current_workbook.value.active, [column_names, statistical_data])
            print(f"Wrote {statistic} data to worksheet {current_workbook.value.active.title!r}.")
        case 3:
            num_to_save = len(unsaved_workbooks)
            if num_to_save == 0:
                print("No workbooks to save.")
                return
            workbooks_to_save = unsaved_workbooks.copy()
            for workbook_name, workbook in workbooks_to_save.items():
                workbook.save(os.path.join("workbooks", f"{workbook_name}.xlsx"))
                del unsaved_workbooks[workbook_name]
            print(f"Successfully saved {num_to_save} workbook(s).")

        case _:
            return


def quit_program_menu():
    message = "Are you sure you want to exit the program? Type 'y' or 'n' to confirm: "
    if unsaved_workbooks:
        message += "You have unsaved workbooks; the data in them will be lost if you do not save them. "
    if get_user_decision(message, "Please confirm whether you would like to exit the program or not."):
        sys.exit()
    else:
        return


if __name__ == "__main__":
    current_workbook = WorkbookSelector(name="selected_workbook", value=None)
    workbook_options = Selector([
        "Create new workbook",
        "Select workbook",
        "Create new worksheet",
        "Select worksheet",
        "Get current workbook name",
        "Get current worksheet name",
        "Rename active worksheet",
    ])
    console_data_options = Selector([
        "Print a column of the Dataset to the console",
        "Print a row of the Dataset to the console",
        "Print the whole Dataset to the console",
        "Print statistical information about the dataset or columns",
        "Print outliers in a column",
        "Print data types of each column",
        "Print every statistical value for a column",
        "Print information about each column",
    ])
    plot_data_selector = Selector([
        "Plot a column against time",
        "Plot two columns on a graph",
    ])
    export_data_selector = Selector([
        "Export a modified Dataset to a spreadsheet",
        "Export a dataset statistic to a spreadsheet",
        "Save all files",
    ])
    statistical_measure_selector = Selector(get_formatted_statistical_function_list())
    main_menu_selection = Selector({"Modify Excel files": spreadsheet_menu,
                                    "Print data from the Dataset to console": data_console_menu,
                                    "Plot data from the Dataset to a graph": data_plotting_menu,
                                    "Export data to Excel files": export_data_menu,
                                    "Quit the program": quit_program_menu})
    print("NOTE: Pressing enter will almost always take you to the previous menu.")
    while main_menu_selection.running:
        main_menu_selection.run()
