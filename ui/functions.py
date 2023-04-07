__all__ = [
    "create_new_directory",
    "get_valid_filename_input",
    "get_valid_worksheet_name",
    "get_workbook_mapping",
    "get_formatted_statistical_function_list",
    "get_statistical_measure_of_region",
    "get_dataset_or_stat_kwargs",
    "get_user_decision",
    "get_valid_column_name",
    "get_valid_row_number",
    "valid_column_name",
    "valid_row_number",
    "validate_argument_value",
    "validate_decimal_places",
    "label_workbooks",
]

import os
from inspect import signature
from typing import Pattern, Callable, Dict, Tuple, List, Literal, Optional, Iterable
from openpyxl import load_workbook

from dataset.constants import FOLDER_DIRECTORY, VALID_NUMERIC_MATCH
from dataset.statmeasures import STATISTICAL_FUNCTIONS  # get_base_statistical_function
from dataset.datasetclass import DatasetStructure
from dataset.datasetclass import Dataset
from ui.constants import VALID_FILENAME, WORKBOOK_DIRECTORY, VALID_EXCEL_SHEET_NAME, VALID_DECISION
from ui.selector import Selector, SelectionDisplay


# from openpyxl.worksheet.worksheet import Worksheet
# from openpxyl.workbook import Workbook

# Private


def _validate_by_regex(input_str: str, invalid_message: str,
                       pattern: Pattern, disallowed_values: list | dict | frozenset = frozenset()) -> str:
    if Selector.PREDETERMINED_INPUT:
        return Selector.next()
    user_input = input(input_str)
    while not pattern.match(user_input) or user_input in disallowed_values:
        print(invalid_message)
        user_input = input(input_str)
    return user_input


def _get_excel_file_names():
    yield from (filename for filename in os.listdir(WORKBOOK_DIRECTORY))


def _join_with_string(join_str: str, end_str: str, iterable: Iterable) -> str:
    items = list(iterable)
    last_item = items.pop(-1)
    return f"{join_str.join(items)} {end_str} {last_item}"


# Public


def create_new_directory(dirname: str):
    os.mkdir(os.path.join(FOLDER_DIRECTORY, dirname))


def andjoin(iterable: Iterable) -> str:
    return _join_with_string(", ", "and", iterable)


def orjoin(iterable: Iterable) -> str:
    return _join_with_string(", ", "or", iterable)


def get_valid_filename_input(existing_filenames: List[str]):
    return _validate_by_regex("Enter a filename for the new excel file: ",
                              "Please re-enter a different filename "
                              "(the filename may already be used or there are invalid characters).",
                              VALID_FILENAME, existing_filenames)


def get_valid_worksheet_name(existing_sheetnames: List[str]):
    return _validate_by_regex("Enter a name for the new worksheet: ",
                              "Please re-enter a different worksheet name ",
                              VALID_EXCEL_SHEET_NAME, existing_sheetnames)


def get_workbook_mapping() -> Dict[str, Callable]:
    # Made a function so that data is updated each time it is called
    excel_files = list(_get_excel_file_names())
    if excel_files:
        return {filename: load_workbook(filename=os.path.join("workbooks", filename)) for filename in excel_files
                if not filename.startswith(".")}
    else:
        return {}

def get_single_argument_functions(function_mapping: dict) -> dict:
    return {function_name: statmeasure.function for function_name, statmeasure in function_mapping.items()
            if len(signature(statmeasure.function).parameters) == 1}


def label_workbooks(unsaved: dict, saved: dict) -> List[Tuple[str, str]]:
    labelled_workbooks = [(workbook_name, "(Unsaved)") for workbook_name in unsaved]
    labelled_workbooks.extend((workbook_name, "(Saved as file)") for workbook_name in saved)
    return sorted(labelled_workbooks)


def get_formatted_statistical_function_list() -> list:
    filtered_functions = get_single_argument_functions(STATISTICAL_FUNCTIONS)
    return sorted(map(lambda function_name: function_name.upper() if len(function_name) <= 3 else function_name.title(),
                      filtered_functions))


def valid_column_name(column_identifier: str, column_names: List[str]) -> str | int | Literal[False]:
    num_of_columns = len(column_names)
    if column_identifier in column_names or \
            (VALID_NUMERIC_MATCH.match(column_identifier) and 0 < int(column_identifier) <= num_of_columns):
        # An interesting property of the above selection: when the regex match is placed first,
        # errors do not occur because the expression returns None, however, if the int(column_identifier)
        # occurs first, then certain errors can still occur (e.g. a string can be inputted and an error raised)
        # In addition, decimals such as 4.0 are invalid input (the user should only enter whole positive numbers)
        return column_names[int(column_identifier) - 1] \
            if column_identifier.isdigit() else column_identifier
    else:
        return False


def get_valid_column_name(column_names: List[str], exclude_first: bool = False) -> Optional[int | str]:
    if exclude_first:
        column_names = column_names[1:]
    print("\nThe columns and their respective number in the dataset are:", SelectionDisplay(column_names))
    column_identifier = input("Enter the name or position of a column in the dataset: ")
    while column_identifier:
        if column_name := valid_column_name(column_identifier, column_names):
            return column_name
        else:
            print("Please re-enter an appropriate column name or column position.")
        column_identifier = input("Enter the name or position of a column in the dataset: ")
    if not column_identifier:
        return None


def get_valid_row_number(num_of_rows: int) -> int:
    row_number = input(f"Enter a row number between 1 and {num_of_rows}: ")
    while row_number:
        if valid_row_number(row_number, num_of_rows):
            return int(row_number)
        else:
            print("Please re-enter an appropriate row number.")
        row_number = input(f"Enter a row number between 1 and {num_of_rows}: ")


def get_statistical_measure_of_region(dataset: Dataset) -> Optional[DatasetStructure]:
    dataset_coverage_selection = Selector(["Column", "Whole Dataset"])
    print()
    print("Which region of the Dataset would you like to be printed?")
    match dataset_coverage_selection.run():
        case 1:
            column_name = get_valid_column_name(dataset.column_names)
            data = dataset[column_name]
        case 2:
            data = dataset
        case _:
            return
    return data


def get_user_decision(input_str: str, invalid_message: str):
    decision = _validate_by_regex(input_str, invalid_message, VALID_DECISION).lower()
    return True if decision.startswith("y") else False


def get_dataset_or_stat_kwargs() -> dict:
    kwargs = {}
    STATISTIC_ARGUMENTS = {
            "na_action": ["remove", "ignore", "average", "mean", "median"],
            "outlier_action": ["remove", "ignore", "average", "mean", "median", "keep"],
        }
    for arg_name, arg_options in STATISTIC_ARGUMENTS.items():
        arg_value = validate_argument_value(
            arg_name,
            f"Enter a value for parameter {arg_name!r} (press enter for default argument): ",
            arg_options,
        )
        if arg_value is not None:
            kwargs[arg_name] = arg_value

    # round decimal
    if (round_dp := validate_decimal_places()) is not None:
        kwargs["round_dp"] = round_dp
    return kwargs


def validate_argument_value(arg_name: str, input_message: str, arg_options: list):
    print(f"The options for the argument {arg_name!r} are: {andjoin(arg_options)}.")
    arg_value = input(input_message) or None
    while arg_value not in arg_options + [None]:
        print("Please re-enter a valid value for the function argument.")
        arg_value = input(input_message) or None
    return arg_value


def valid_row_number(row_number: str, num_of_rows: int) -> bool:
    return VALID_NUMERIC_MATCH.match(row_number) and 0 < int(row_number) <= num_of_rows


def validate_decimal_places() -> Optional[int]:
    input_message = "How many decimal places would you like to round the value to (press enter for no rounding): "
    decimal_places = input(input_message)
    while decimal_places and not VALID_NUMERIC_MATCH.match(decimal_places):
        print("Please re-enter a valid integer.")
        decimal_places = input(input_message)
    return None if not decimal_places else int(decimal_places)
