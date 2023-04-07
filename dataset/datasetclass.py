__all__ = ["Dataset", "DatasetStructure"]

import os
from datetime import timedelta
from inspect import signature
from operator import attrgetter, itemgetter
from typing import Any, Callable, List, Tuple, Generator
from collections import Counter

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from dataset.constants import DATA_FILE_DIRECTORY, EXCEL_FILE_NAME
from dataset.functions import (
    _bound_worksheet_data_region, _generate_structure_string, _column_number_to_letter, _flatten,
    _get_cell_values, _date_string_to_datetime, _replace_nones, _format_slice, _remove_nans, _get_array_dtype
)
from dataset.statmeasures import STATISTICAL_FUNCTIONS, Numeric, get_base_statistical_function, reformat_data
from dataset.structures import _DatasetArrayRow, _DatasetArrayColumnView, _DatasetArray, _Schema
from dataset.config import indentation_character, dataset_configurables


def _generate_array_from_worksheet(worksheet: Worksheet) -> Tuple[list, Any]:
    index_row, last_data_row, last_data_column = _bound_worksheet_data_region(worksheet)
    blank_column_numbers = []
    column_names = []
    for column_number in range(1, last_data_column + 1):
        column_letter = _column_number_to_letter(column_number)
        cell_range = _flatten(worksheet[f"{column_letter}{index_row}:{column_letter}{last_data_row}"])
        index_name, *column_values = _get_cell_values(cell_range)
        if all(value is None for value in column_values):
            blank_column_numbers.append(column_number)
        else:
            column_names.append(index_name)

    array = []
    for row_number in range(index_row + 1, last_data_row + 1):
        cell_range = _flatten(worksheet[f"A{row_number}:{_column_number_to_letter(last_data_column)}{row_number}"])
        cell_value_column_mapping = list(enumerate(_get_cell_values(cell_range), 1))
        for index, (column_number, _) in enumerate(cell_value_column_mapping):
            if column_number in blank_column_numbers:
                del cell_value_column_mapping[index]

        array.append(_DatasetArrayRow(list(map(itemgetter(1), cell_value_column_mapping)), column_names))

    return column_names, _DatasetArray(array)


class _Dataset:

    def __init__(self, excel_file_name: str):
        self._workbook_name = excel_file_name
        self._workbook = load_workbook(os.path.join(DATA_FILE_DIRECTORY, excel_file_name))
        self._worksheet = self._workbook.active
        self._column_names, self._array = _generate_array_from_worksheet(self._worksheet)
        self._schema = self._generate_dataset_schema()
        self._apply_function_to_column("Date", _date_string_to_datetime)
        for row_index in range(len(self._array)):
            self._apply_function_to_row(row_index, _replace_nones)
        for column_name in self._column_names:
            self._cast_to_one_type(column_name)

    def __iter__(self):
        return iter(self._array)

    def __getitem__(self, index: int | str) -> _DatasetArrayRow | _DatasetArrayColumnView:
        if type(index) is str:
            # Access a column
            return self._get_column_view(index)
        elif type(index) is int:
            # Access a row
            return self._array[index]

    def __setitem__(self, column_name: str, column: _DatasetArrayColumnView):
        column_number = self._column_names.index(column_name)
        for index, value in enumerate(column):
            self._array[index][column_number] = value

    def __len__(self):
        return len(self._array)

    def __str__(self) -> str:
        ellipsis_space = (2 * 2 + 3) * " "  # 2 * space around + 1 (elipsis length)
        min_column_characters = max(map(len, self._column_names))

        length = len(self._column_names)
        characters, left, right = _format_slice(min_column_characters + 2, len(ellipsis_space),
                                                length)  # + 2 for readability
        # Within f strings, "{{" means a literal "{" (since characters in f-strings cannot be escaped
        column_format = f"{{:{indentation_character.value}{characters}}}"
        row_format = column_format * left \
                     + ellipsis_space \
                     + "..." + ellipsis_space \
                     + column_format * ((length - right) - 2)
        output = row_format.format(*self._column_names[0:left], *self._column_names[right + 1:length])
        for row in self._array:
            output += "\n" + row_format.format(*(map(str, row[0:left] + row[right + 1:length])))
        return output

    # Noteworthy point: properties defined here have only getters (unlike the public attributes which have setters too)

    @property
    def column_names(self) -> List[str]:
        return self._column_names

    @property
    def schema(self) -> _Schema:
        return self._schema

    def _cast_to_one_type(self, column_name: str):
        column_data = self._get_column_data(column_name, remove_nans=True)
        dtype_counter = Counter(map(type, column_data))
        dtype_occurences = list(dtype_counter.items())
        # Reverse sort so that the most frequently appearing data type is first in the list
        dtype_occurences.sort(key=itemgetter(1), reverse=True)
        if not len(dtype_occurences) == 1:
            # The first data type (i.e. most occuring) is indexed and cast to the whole column
            cast_type = dtype_occurences[0][0]
            self._apply_function_to_column(column_name, cast_type)

    def _get_column_dtype(self, column_name: str) -> type:
        column_data = self._get_column_data(column_name, remove_nans=True)
        return _get_array_dtype(list(column_data))

    def _get_column_data(self, column_name: str, remove_nans: bool = False) -> list:
        column_index = self._column_names.index(column_name)
        column_data = [self._array[i][column_index] for i in range(len(self._array))]
        return column_data if not remove_nans else _remove_nans(column_data)

    def _get_column_view(self, column_name: str) -> _DatasetArrayColumnView:
        # Implemented to remove recursive calls between the above two functions
        # Also, this is much more readable with each function call returning its respective class attribute
        return _DatasetArrayColumnView(column_name,
                                       self._get_column_data(column_name),
                                       self._get_column_dtype(column_name)
                                       )

    def _generate_dataset_schema(self) -> _Schema:
        _, last_data_row, last_data_column = _bound_worksheet_data_region(self._worksheet)

        blank_column_numbers = []
        for column_number in range(1, last_data_column + 1, 1):
            column_letter = _column_number_to_letter(column_number)
            cell_range = _flatten(self._worksheet[f"{column_letter}1:{column_letter}{last_data_row}"])
            cell_values = _get_cell_values(cell_range)
            if all(cell_value is None for cell_value in cell_values):
                blank_column_numbers.append(column_number)

        # In the current dataset, the statement below is true
        # assert blank_column_numbers == [3, 5]

        top_cells = ([], [])
        for i, row in enumerate(self._worksheet.iter_rows(min_row=1, min_col=1, max_row=2, max_col=last_data_column)):
            top_cells[i].extend(_get_cell_values(_flatten(row)))

        column_names = []
        column_titles = []
        last_valid_column_title = None
        for index in range(len(top_cells[0])):
            # Excel index (i.e. columns) starts at 1; Python starts at 0 for indexing the list
            if index + 1 not in blank_column_numbers:
                column_title = top_cells[0][index]
                column_name = top_cells[1][index]
                if column_title is not None:
                    last_valid_column_title = column_title
                if last_valid_column_title is not None:
                    column_names.append(column_name)
                    column_titles.append(last_valid_column_title)

        return _Schema(dict(zip(column_names, column_titles)))

    def _apply_function_to_column(self, column_name: str, function: Callable):
        column_position = self._column_names.index(column_name)
        for index in range(len(self[column_name])):
            self._array[index].apply_function_at_index(function, column_position)

    def _apply_function_to_row(self, row_index: int, function: Callable):
        self._array[row_index].apply_function(function)

    def _column_iterator(self) -> Generator[_DatasetArrayColumnView, Any, None]:
        yield from map(self._get_column_view, self._column_names)


class Dataset(_Dataset):

    def __init__(self, excel_file_name: str, dataset_name: str):
        super().__init__(excel_file_name)
        self.dataset_name = dataset_name

    def get_column_dtype(self, column_name: str, type_string: bool = True) -> type | str:
        dtype = self._get_column_dtype(column_name)
        return dtype.__name__ if type_string else dtype

    def get_column_dtypes(self) -> str:
        return _generate_structure_string([self._column_names,
                                           list(map(self.get_column_dtype, self._column_names))],
                                          ["Column", "Data type"])

    def get_timedeltas_between_measurements(self) -> List[timedelta]:
        column_length = len(date_column := self["Date"])
        differences = []
        for index in range(column_length - 1):
            value, next_value = date_column[index], date_column[index + 1]
            differences.append(next_value - value)
        return differences

    def get_average_timedelta(self):
        timedeltas = self.get_timedeltas_between_measurements()
        return timedelta(days=sum(map(attrgetter("days"), timedeltas)) // len(timedeltas))

    def get_stat_of_columns(self, statistic: str, *args: Any, **kwargs: Any) -> List[Numeric | None]:
        return [STATISTICAL_FUNCTIONS[statistic](column, *args, **kwargs) for column in self._column_iterator()]

    def get_outliers_in_column(self, column_name: str) -> Tuple[tuple, tuple]:
        outlier_function = get_base_statistical_function("outlier")
        filtered_data = self.filter_column(column_name, outlier_function)
        if not filtered_data:
            outlier_row_indexes, outlier_values = [], []
        else:
            outlier_row_indexes, outlier_values = zip(*filtered_data)
        return outlier_row_indexes, outlier_values

    def get_outlier_string(self, column_name: str) -> str:
        values = self.get_outliers_in_column(column_name)
        if not values[0]:
            return ""
        output = _generate_structure_string(list(values), ["Index", column_name], cut_data=False)
        output += f"\n\nMean: {self[column_name].statistic('mean')}\n" \
                  f"Q1: {self[column_name].statistic('q1')}\n" \
                  f"Median (Q2): {self[column_name].statistic('median')}\n" \
                  f"Q3: {self[column_name].statistic('q3')}\n" \
                  f"IQR: {self[column_name].statistic('iqr')}\n"
        return output

    @staticmethod
    def get_statistic_string(identifier_column: list, data_column: list, statistic: str, cut_data: bool = False):
        return _generate_structure_string([identifier_column, data_column],
                                          ["Column Name", statistic.replace("_", " ").capitalize()],
                                          cut_data=cut_data)

    @staticmethod
    def set_config(config_name, value):
        # Case sensitive
        for configurable in dataset_configurables:
            if configurable.name == config_name:
                configurable.register_value(value)
                return
        else:
            raise Exception("Configuration not found")

    def statistic(self, statistic: str, *args: Any, **kwargs: Any) -> Tuple[list, list]:
        colnames_column = self._column_names
        data_column = self.get_stat_of_columns(statistic, *args, **kwargs)
        return colnames_column, data_column

    def filter_column(self, column_name: str, function: Callable) -> list:
        data = self[column_name]
        filter_matches = []
        function_argcount = len(signature(function).parameters)
        if function_argcount == 2:
            for index, value in enumerate(data):
                if function(data, value):
                    filter_matches.append((index, value))
        elif function_argcount == 1:
            pass
        return filter_matches

    def reformat(self, *, na_action: str = "ignore", outlier_action: str = "keep", round_dp: bool = False):
        reformatted_dataset = Dataset(self._workbook_name, self.dataset_name)
        for index in range(len(self._column_names), 1):
            column_name = self._column_names[index]
            column = self[column_name]
            data = reformat_data(list(column), na_action=na_action, outlier_action=outlier_action)
            if round_dp:
                data = [round(value, round_dp) for value in data]
            reformatted_dataset[column_name] = data
        return reformatted_dataset


DatasetStructure = Dataset | _DatasetArrayColumnView | _DatasetArrayRow

if __name__ == "__main__":
    dataset = Dataset(EXCEL_FILE_NAME, "Logan's Dam Water Quality")
    # print(dataset.statistic("mean", round_dp=5, outlier_action="ignore"))
    # print(dataset.filter_column(dataset.column_names[1], STATISTICAL_FUNCTIONS["outlier"].function))
    print(dataset.get_outlier_string(dataset.column_names[5]))
    # print(dataset.get_stat_of_columns("mean", round_dp=5, outlier_action="average"))
