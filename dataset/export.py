__all__ = [
    "write_dataset_to_worksheet",
    "write_dataset_column_to_worksheet",
    "write_columns_to_worksheet",
]

from typing import Iterable

from openpyxl.styles import Border, Side
from openpyxl.workbook import Workbook
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from dataset.functions import _column_number_to_letter, _datetime_to_date_string, _column_letter_to_number
from dataset.datasetclass import Dataset

THICK_BOTTOM_BORDER = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thick'))


def write_dataset_to_worksheet(worksheet: Worksheet, ds: Dataset):
    schema = ds.schema
    previous_title = None
    write_dataset_column_to_worksheet(worksheet, "A", list(map(_datetime_to_date_string, ds[ds.column_names[0]])))
    for index in range(1, len(ds.column_names)):
        column_data = []
        column_name, column_letter = ds.column_names[index], _column_number_to_letter(index)
        if (title := schema[column_name]) != previous_title:
            previous_title = title
            column_data.append(title)
        else:
            column_data.append(None)
        column_data.append(column_name)
        column_data.extend(list(ds[column_name]))
        write_dataset_column_to_worksheet(worksheet, column_letter, column_data)


def write_dataset_column_to_worksheet(worksheet: Worksheet, column_letter: str,
                                      column_data: Iterable, add_border: bool = True):
    column_number = _column_letter_to_number(column_letter)
    for index, value in enumerate(column_data, 1):
        worksheet.cell(row=index, column=column_number).value = value
        if add_border and index == 2:
            worksheet.cell(row=index, column=column_number).border = THICK_BOTTOM_BORDER
    if add_border:
        worksheet.cell(row=index, column=column_number).border = THICK_BOTTOM_BORDER


def write_columns_to_worksheet(worksheet: Worksheet, columns: list):
    longest_length = max(len(column) for column in columns)
    for column_number in range(1, len(columns) + 1):
        column = columns[column_number - 1]
        for row_number in range(1, longest_length + 1):
            if row_number >= len(column):
                value = None
            else:
                value = str(column[row_number - 1])
            worksheet.cell(row=row_number, column=column_number).value = value

